import csv
import json
import sys
import tempfile
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import sheet_importer


class SheetImporterTests(unittest.TestCase):
    def test_csv_supports_forward_filled_city_and_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir, "pois.csv")
            path.write_text(
                "目的地,编号,地标\nSeoul,2,Gyeongbokgung Palace\n,1,Namsan Tower\n",
                encoding="utf-8",
            )
            city, pois = sheet_importer.extract_pois(path, "Seoul")
        self.assertEqual(city, "Seoul")
        self.assertEqual(pois, ["Namsan Tower", "Gyeongbokgung Palace"])

    def test_xlsx_auto_detects_header_sheet(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir, "pois.xlsx")
            workbook = Workbook()
            ws = workbook.active
            ws.title = "导览卡图标"
            ws.append(["说明"])
            ws.append(["目的地", "编号", "地标"])
            ws.append(["Seoul", 1, "Namsan Tower"])
            workbook.save(path)
            city, pois = sheet_importer.extract_pois(path, "Seoul")
        self.assertEqual(city, "Seoul")
        self.assertEqual(pois, ["Namsan Tower"])

    def test_extract_all_cities_preserves_descriptions(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir, "all.csv")
            path.write_text(
                "城市,序号,地标,中文,视觉描述\nSeoul,1,Namsan Tower,南山首尔塔,White observation tower\n"
                "Tokyo,1,Tokyo Tower,东京塔,Red lattice tower\n",
                encoding="utf-8",
            )
            grouped = sheet_importer.extract_all_cities(path)
        self.assertEqual(list(grouped), ["Seoul", "Tokyo"])
        self.assertEqual(grouped["Seoul"][0]["description"], "White observation tower")
        self.assertEqual(grouped["Seoul"][0]["name_zh"], "南山首尔塔")

    def test_full_xlsx_chooses_sheet_with_most_pois(self):
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir, "multi_sheet.xlsx")
            workbook = Workbook()
            first = workbook.active
            first.title = "Summary"
            first.append(["城市", "地标"])
            first.append(["Seoul", "One"])
            detailed = workbook.create_sheet("导览卡图标")
            detailed.append(["城市", "编号", "地标"])
            detailed.append(["Seoul", 1, "One"])
            detailed.append([None, 2, "Two"])
            detailed.append([None, 3, "Three"])
            workbook.save(path)
            grouped = sheet_importer.extract_all_cities(path)
        self.assertEqual([item["name"] for item in grouped["Seoul"]], ["One", "Two", "Three"])


if __name__ == "__main__":
    unittest.main()
