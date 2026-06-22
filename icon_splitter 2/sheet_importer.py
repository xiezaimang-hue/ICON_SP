#!/usr/bin/env python3
"""Spreadsheet parsing helpers for manual city asset imports."""

from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


CITY_COLUMNS = {"city", "城市", "destination", "目的地", "place", "地点"}
POI_COLUMNS = {"poi", "景点", "地标", "attraction", "name", "名称"}
ORDER_COLUMNS = {"order", "序号", "编号", "index", "idx", "排序", "position"}
DESCRIPTION_COLUMNS = {
    "description", "visual description", "icon description",
    "描述", "视觉描述", "图标描述", "内容描述",
}
CHINESE_COLUMNS = {"中文", "中文名", "中文名称", "chinese", "name_zh", "zh_name"}


def normalize_text(value, *, newline_replacement=" ") -> str:
    if value is None:
        return ""
    text = re.sub(r"\s*[\r\n]+\s*", newline_replacement, str(value).strip())
    return re.sub(r"[ \t]+", " ", text).strip()


def pick_column(headers: Iterable[str], candidates: set[str]) -> Optional[str]:
    original = [str(x).strip().replace("\ufeff", "") for x in headers if x is not None]
    lowered = {x.lower(): x for x in original}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    return None


def parse_records(headers: Sequence[str], rows: Iterable[Dict[str, object]], city_hint: str) -> Dict[str, List[Tuple[int, str]]]:
    poi_col = pick_column(headers, POI_COLUMNS)
    city_col = pick_column(headers, CITY_COLUMNS)
    order_col = pick_column(headers, ORDER_COLUMNS)
    if not poi_col:
        raise ValueError("表格中找不到 POI/景点/地标 列")
    grouped: Dict[str, List[Tuple[int, str]]] = defaultdict(list)
    current_city = normalize_text(city_hint, newline_replacement=" / ")
    fallback = 0
    for row in rows:
        city = normalize_text(row.get(city_col), newline_replacement=" / ") if city_col else current_city
        if city:
            current_city = city
        poi = normalize_text(row.get(poi_col))
        if not current_city or not poi:
            continue
        fallback += 1
        order = fallback
        if order_col:
            raw = normalize_text(row.get(order_col))
            if raw:
                try:
                    order = int(float(raw))
                except ValueError:
                    pass
        grouped[current_city].append((order, poi))
    for city in grouped:
        grouped[city].sort(key=lambda item: item[0])
    return grouped


def parse_spec_records(
    headers: Sequence[str],
    rows: Iterable[Dict[str, object]],
    city_hint: str = "",
) -> Dict[str, List[Tuple[int, dict]]]:
    """Parse all cities while preserving optional per-POI visual descriptions."""
    poi_col = pick_column(headers, POI_COLUMNS)
    city_col = pick_column(headers, CITY_COLUMNS)
    order_col = pick_column(headers, ORDER_COLUMNS)
    description_col = pick_column(headers, DESCRIPTION_COLUMNS)
    chinese_col = pick_column(headers, CHINESE_COLUMNS)
    if not poi_col:
        raise ValueError("表格中找不到 POI/景点/地标 列")
    if not city_col and not city_hint:
        raise ValueError("整表导入需要 city/城市/目的地 列")

    grouped: Dict[str, List[Tuple[int, dict]]] = defaultdict(list)
    current_city = normalize_text(city_hint, newline_replacement=" / ")
    fallback = 0
    for row in rows:
        city = normalize_text(row.get(city_col), newline_replacement=" / ") if city_col else current_city
        if city:
            current_city = city
        poi = normalize_text(row.get(poi_col))
        if not current_city or not poi:
            continue
        fallback += 1
        order = fallback
        if order_col:
            raw_order = normalize_text(row.get(order_col))
            if raw_order:
                try:
                    order = int(float(raw_order))
                except ValueError:
                    pass
        description = normalize_text(row.get(description_col)) if description_col else ""
        name_zh = normalize_text(row.get(chinese_col)) if chinese_col else ""
        grouped[current_city].append((order, {
            "name": poi,
            "name_zh": name_zh,
            "description": description,
        }))
    for city in grouped:
        grouped[city].sort(key=lambda item: item[0])
    return grouped


def _select_city(grouped: Dict[str, List[Tuple[int, str]]], city_hint: str) -> Tuple[str, List[str]]:
    if not grouped:
        raise ValueError("表格中没有读取到任何 POI")
    hint = normalize_text(city_hint, newline_replacement=" / ").casefold()
    exact = [city for city in grouped if city.casefold() == hint]
    if not exact:
        exact = [city for city in grouped if hint in city.casefold() or city.casefold() in hint]
    if exact:
        selected = exact[0]
    elif len(grouped) == 1:
        selected = next(iter(grouped))
    else:
        raise ValueError("表格包含多个城市，但未找到与输入城市名匹配的数据：" + "、".join(grouped))
    return selected, [poi for _, poi in grouped[selected]]


def parse_csv(path: Path, city_hint: str) -> Tuple[str, List[str]]:
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("无法识别 CSV 文件编码")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV 中没有表头")
    return _select_city(parse_records(reader.fieldnames, reader, city_hint), city_hint)


def _row_values(ws, row_index: int):
    return [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]


def parse_xlsx(path: Path, city_hint: str) -> Tuple[str, List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法读取 XLSX") from exc
    workbook = load_workbook(path, read_only=False, data_only=True)
    last_error = None
    for ws in workbook.worksheets:
        for header_row in range(1, min(ws.max_row, 30) + 1):
            values = _row_values(ws, header_row)
            headers = ["" if value is None else str(value).strip() for value in values]
            if not pick_column(headers, POI_COLUMNS):
                continue
            records = []
            for row_index in range(header_row + 1, ws.max_row + 1):
                row = _row_values(ws, row_index)
                records.append({headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]})
            try:
                grouped = parse_records(headers, records, city_hint)
                if grouped:
                    return _select_city(grouped, city_hint)
            except Exception as exc:
                last_error = exc
                break
    if last_error:
        raise last_error
    raise ValueError("找不到包含 POI/景点/地标 列的工作表")


def extract_pois(path: Path, city_hint: str) -> Tuple[str, List[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return parse_xlsx(path, city_hint)
    if suffix == ".csv":
        return parse_csv(path, city_hint)
    raise ValueError("表格只支持 .xlsx 或 .csv")


def extract_all_cities(path: Path) -> Dict[str, List[dict]]:
    """Read one XLSX/CSV and return every city as ordered POI specs."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        raw = path.read_bytes()
        text = None
        for encoding in ("utf-8-sig", "gb18030", "utf-16"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("无法识别 CSV 文件编码")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise ValueError("CSV 中没有表头")
        grouped = parse_spec_records(reader.fieldnames, reader)
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl，无法读取 XLSX") from exc
        workbook = load_workbook(path, read_only=False, data_only=True)
        grouped = {}
        best_total = 0
        last_error = None
        for ws in workbook.worksheets:
            for header_row in range(1, min(ws.max_row, 30) + 1):
                values = _row_values(ws, header_row)
                headers = ["" if value is None else str(value).strip() for value in values]
                if not pick_column(headers, POI_COLUMNS) or not pick_column(headers, CITY_COLUMNS):
                    continue
                records = []
                for row_index in range(header_row + 1, ws.max_row + 1):
                    row = _row_values(ws, row_index)
                    records.append({headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]})
                try:
                    candidate = parse_spec_records(headers, records)
                    if candidate:
                        candidate_total = sum(len(items) for items in candidate.values())
                        if candidate_total > best_total:
                            grouped = candidate
                            best_total = candidate_total
                        break
                except Exception as exc:
                    last_error = exc
        if not grouped:
            if last_error:
                raise last_error
            raise ValueError("找不到同时包含城市和POI列的工作表")
    else:
        raise ValueError("表格只支持 .xlsx 或 .csv")

    return {city: [spec for _, spec in ordered] for city, ordered in grouped.items()}


def natural_key(filename: str):
    return [int(x) if x.isdigit() else x.casefold() for x in re.split(r"(\d+)", filename)]
