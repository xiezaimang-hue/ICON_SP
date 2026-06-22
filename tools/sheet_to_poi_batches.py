#!/usr/bin/env python3
"""
Extract city + POI rows from a CSV file or CSV-exportable online sheet,
then split POIs into 16-item prompt pages.

Supported input:
- Local XLSX file
- Local CSV file
- Direct CSV URL
- Google Sheets share URL, if the sheet is public or published

Expected columns:
- city / 城市 / destination / place
- poi / POI / 景点 / 地标 / name
- optional order / 序号 / index
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


CITY_COLUMNS = {"city", "城市", "destination", "目的地", "place", "地点"}
POI_COLUMNS = {"poi", "POI", "景点", "地标", "attraction", "name", "名称"}
ORDER_COLUMNS = {"order", "序号", "index", "idx", "排序", "position"}


def normalize_header(value: str) -> str:
    return value.strip().replace("\ufeff", "")


def pick_column(headers: Iterable[str], candidates: set[str]) -> Optional[str]:
    normalized = {normalize_header(h): h for h in headers}
    lowered = {normalize_header(h).lower(): h for h in headers}
    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    return None


def normalize_cell_text(value: object, *, newline_replacement: str = " ") -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s*[\r\n]+\s*", newline_replacement, text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def google_sheet_to_csv_url(url: str) -> str:
    """Convert a regular Google Sheets URL to a CSV export URL when possible."""
    parsed = urllib.parse.urlparse(url)
    if "docs.google.com" not in parsed.netloc or "/spreadsheets/" not in parsed.path:
        return url

    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        return url

    sheet_id = match.group(1)
    query = urllib.parse.parse_qs(parsed.query)
    gid = query.get("gid", ["0"])[0]

    if parsed.path.endswith("/pub"):
        return url

    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def read_source(source: str) -> str:
    if re.match(r"^https?://", source):
        csv_url = google_sheet_to_csv_url(source)
        req = urllib.request.Request(csv_url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as response:
            content_type = response.headers.get("content-type", "")
            text = response.read().decode("utf-8-sig")
            looks_like_html = text.lstrip().lower().startswith(("<!doctype html", "<html"))
            if "text/html" in content_type.lower() or looks_like_html:
                lowered = text.lower()
                if "login" in lowered or "oauth" in lowered or "身份认证" in text:
                    raise PermissionError(
                        "The URL returned a login/authentication page, not sheet data. "
                        "Export or publish the sheet as CSV, or provide a public CSV URL."
                    )
                raise ValueError(
                    "The URL returned HTML, not CSV data. "
                    "Use a direct CSV export URL or a published public sheet."
                )
            return text

    path = Path(source)
    if not path.exists():
        raise FileNotFoundError(f"Source not found: {source}")
    return path.read_text(encoding="utf-8-sig")


def parse_records(fieldnames: Sequence[str], rows: Iterable[Dict[str, object]]) -> Dict[str, List[Tuple[int, str]]]:
    if not fieldnames:
        raise ValueError("No header row found")

    city_col = pick_column(fieldnames, CITY_COLUMNS)
    poi_col = pick_column(fieldnames, POI_COLUMNS)
    order_col = pick_column(fieldnames, ORDER_COLUMNS)

    if not city_col or not poi_col:
        raise ValueError(
            "Missing required columns. Need city/城市 and poi/景点 columns. "
            f"Found columns: {', '.join(str(h) for h in fieldnames)}"
        )

    grouped: Dict[str, List[Tuple[int, str]]] = defaultdict(list)
    fallback_order = 0
    current_city = ""

    for row in rows:
        raw_city = row.get(city_col)
        raw_poi = row.get(poi_col)
        city = normalize_cell_text(raw_city, newline_replacement=" / ")
        poi = normalize_cell_text(raw_poi)

        if city:
            current_city = city
        elif current_city:
            city = current_city

        if not city or not poi:
            continue

        fallback_order += 1
        order = fallback_order
        if order_col:
            raw_order_value = row.get(order_col)
            raw_order = "" if raw_order_value is None else str(raw_order_value).strip()
            if raw_order:
                try:
                    order = int(float(raw_order))
                except ValueError:
                    order = fallback_order

        grouped[city].append((order, poi))

    for city in grouped:
        grouped[city].sort(key=lambda item: item[0])

    return grouped


def parse_csv_rows(csv_text: str) -> Dict[str, List[Tuple[int, str]]]:
    reader = csv.DictReader(io.StringIO(csv_text))
    if not reader.fieldnames:
        raise ValueError("CSV has no header row")
    return parse_records(reader.fieldnames, reader)


def row_values(ws, row_index: int) -> List[object]:
    return [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]


def find_header_row(ws) -> Optional[Tuple[int, List[str]]]:
    max_scan = min(ws.max_row, 30)
    for row_index in range(1, max_scan + 1):
        values = row_values(ws, row_index)
        headers = ["" if value is None else str(value).strip() for value in values]
        if pick_column(headers, CITY_COLUMNS) and pick_column(headers, POI_COLUMNS):
            return row_index, headers
    return None


def parse_xlsx_rows(source: str, preferred_sheet: Optional[str] = None) -> Dict[str, List[Tuple[int, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read .xlsx files") from exc

    wb = load_workbook(source, read_only=False, data_only=True)
    worksheets = wb.worksheets
    if preferred_sheet:
        if preferred_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {preferred_sheet}. Available sheets: {', '.join(wb.sheetnames)}")
        worksheets = [wb[preferred_sheet]]

    last_error: Optional[Exception] = None
    for ws in worksheets:
        header = find_header_row(ws)
        if not header:
            continue
        header_row, headers = header
        records = []
        for row_index in range(header_row + 1, ws.max_row + 1):
            values = row_values(ws, row_index)
            record = {
                headers[col_index]: values[col_index]
                for col_index in range(min(len(headers), len(values)))
                if headers[col_index]
            }
            records.append(record)
        try:
            grouped = parse_records(headers, records)
        except Exception as exc:
            last_error = exc
            continue
        if grouped:
            return grouped

    if last_error:
        raise last_error
    raise ValueError(
        "Could not find a worksheet with city/城市 and poi/景点 columns. "
        f"Available sheets: {', '.join(wb.sheetnames)}"
    )


def chunked(items: List[str], size: int = 16) -> List[List[str]]:
    return [items[i:i + size] for i in range(0, len(items), size)]


def safe_slug(value: str) -> str:
    slug = re.sub(r"[^\w\-]+", "_", value.strip(), flags=re.UNICODE).strip("_")
    return slug or "city"


def build_outputs(grouped: Dict[str, List[Tuple[int, str]]]) -> dict:
    cities = []
    for city, ordered_pois in grouped.items():
        pois = [poi for _, poi in ordered_pois]
        pages = chunked(pois, 16)
        cities.append({
            "city": city,
            "total_pois": len(pois),
            "pages": [
                {
                    "page": index + 1,
                    "start_index": index * 16 + 1,
                    "end_index": index * 16 + len(page),
                    "poi_count": len(page),
                    "pois": page,
                }
                for index, page in enumerate(pages)
            ],
        })
    return {"cities": cities}


def write_agent_requests(data: dict, output_path: Path) -> None:
    lines = [
        "# POI Prompt Requests",
        "",
        "Use `agent_rules/poi_icon_prompt_rule.md` to generate prompts for each page below.",
        "",
    ]

    for city_data in data["cities"]:
        city = city_data["city"]
        lines.append(f"## {city}")
        lines.append("")
        for page in city_data["pages"]:
            lines.append(f"### PAGE {page['page']} ({page['poi_count']} POIs)")
            lines.append("")
            lines.append(f"城市：{city}")
            lines.append("POI：")
            for idx, poi in enumerate(page["pois"], start=1):
                lines.append(f"{idx}. {poi}")
            lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract city + POI rows from an XLSX/CSV/online sheet and split them into 16-item pages."
    )
    parser.add_argument("source", help="Local XLSX/CSV path, direct CSV URL, or public Google Sheets URL")
    parser.add_argument("--sheet", help="Worksheet name for XLSX input. If omitted, auto-detects a sheet.")
    parser.add_argument(
        "--out-dir",
        default="generated_prompts",
        help="Output directory for extracted JSON and agent request markdown",
    )
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    source_path = Path(args.source)
    if source_path.suffix.lower() == ".xlsx" and source_path.exists():
        grouped = parse_xlsx_rows(args.source, args.sheet)
    else:
        csv_text = read_source(args.source)
        grouped = parse_csv_rows(csv_text)
    data = build_outputs(grouped)

    json_path = out_dir / "poi_batches.json"
    md_path = out_dir / "agent_prompt_requests.md"

    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    write_agent_requests(data, md_path)

    print(f"[OK] Cities: {len(data['cities'])}")
    for city_data in data["cities"]:
        print(
            f"  - {city_data['city']}: {city_data['total_pois']} POIs, "
            f"{len(city_data['pages'])} page(s)"
        )
    print(f"[OK] Wrote {json_path}")
    print(f"[OK] Wrote {md_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1)
